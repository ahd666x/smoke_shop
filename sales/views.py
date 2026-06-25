import json
import io
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta, date

from django.db.models.functions import TruncDate, TruncHour
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.db import transaction, models
from django.db.models import Sum, F, Q, Count, Prefetch, DecimalField
from django.utils import timezone
from django.contrib import messages
from django.apps import apps

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Order, OrderItem, Payment, Customer
from inventory.models import Product, StockMovement, Purchase, PurchaseItem, Supplier, Category
from taxation.services import calculate_item_tax
from taxation.models import TaxLine, ProductTaxProfile

import jdatetime

# ==================== ویوهای کمکی ====================

def _parse_qty(value):
    qty = Decimal(str(value))
    if qty <= 0:
        raise ValueError('invalid quantity')
    return qty


def _get_avg_cost(product):
    """محاسبه میانگین موزون بهای تمام‌شده"""
    if product.stock_quantity > 0:
        return (product.inventory_cost / product.stock_quantity).quantize(
            Decimal('1'), rounding=ROUND_HALF_UP
        )
    return product.purchase_price


def recalc_cart(request):
    """بازگرداندن خلاصه سبد با احتساب تخفیف"""
    cart = request.session.get('cart', [])
    discount = request.session.get('discount', {'type': 'none', 'value': 0})

    product_ids = [item['product_id'] for item in cart]
    products_map = {p.pk: p for p in Product.objects.filter(pk__in=product_ids)}

    subtotal = Decimal('0')
    total_tax = Decimal('0')

    for item in cart:
        product = products_map.get(item['product_id'])
        if not product:
            continue
        qty = _parse_qty(item['quantity'])
        price = Decimal(str(item['unit_price']))
        tax, _ = calculate_item_tax(price, qty, product)
        item['tax'] = float(tax)
        subtotal += price * qty
        total_tax += tax
        if 'purchase_cost' not in item:
            item['purchase_cost'] = 0.0

    before_discount = subtotal + total_tax
    discount_amount = Decimal('0')
    if discount['type'] == 'percent':
        discount_amount = (before_discount * Decimal(str(discount['value']))) / Decimal('100')
    elif discount['type'] == 'amount':
        discount_amount = Decimal(str(discount['value']))
    grand_total = before_discount - discount_amount

    cart_response = [{
        'product_id': it['product_id'],
        'name': it['name'],
        'barcode': it['barcode'],
        'unit_price': it['unit_price'],
        'quantity': it['quantity'],
        'tax': it['tax'],
        'purchase_cost': it.get('purchase_cost', 0.0)
    } for it in cart]

    return JsonResponse({
        'cart': cart_response,
        'subtotal': float(subtotal),
        'total_tax': float(total_tax),
        'before_discount': float(before_discount),
        'discount_type': discount['type'],
        'discount_value': float(discount['value']),
        'discount_amount': float(discount_amount),
        'grand_total': float(grand_total),
    })


# ==================== ویوهای صندوق فروش (POS) ====================

@login_required
def pos_view(request):
    request.session['cart'] = []
    request.session['discount'] = {'type': 'none', 'value': 0}
    customers = Customer.objects.all()[:50]
    held_count = Order.objects.filter(is_held=True).count()
    return render(request, 'sales/pos.html', {
        'customers': customers,
        'held_count': held_count,
    })


@login_required
def search_products(request):
    query = request.GET.get('q', '')
    products = Product.objects.filter(is_active=True).filter(
        Q(name__icontains=query) | Q(barcode__icontains=query)
    )[:10]
    results = []
    for p in products:
        cost = float(_get_avg_cost(p))
        results.append({
            'id': p.id,
            'name': str(p),
            'barcode': p.barcode,
            'price': float(p.selling_price),
            'cost': cost,
            'stock': float(p.stock_quantity),
        })
    return JsonResponse(results, safe=False)


@login_required
def get_featured_products(request):
    products = Product.objects.filter(is_active=True, is_featured=True)[:12]
    data = [{'id': p.id, 'name': str(p), 'barcode': p.barcode, 'price': float(p.selling_price)} for p in products]
    return JsonResponse(data, safe=False)


@login_required
@require_POST
def add_to_cart(request):
    data = json.loads(request.body)
    barcode = data.get('barcode')
    product_id = data.get('product_id')
    if barcode:
        product = get_object_or_404(Product, barcode=barcode, is_active=True)
    elif product_id:
        product = get_object_or_404(Product, pk=product_id, is_active=True)
    else:
        return JsonResponse({'error': 'بارکد یا شناسه محصول الزامی است'}, status=400)

    purchase_cost = float(_get_avg_cost(product))

    cart = request.session.get('cart', [])
    found = False
    for item in cart:
        if item['product_id'] == product.id:
            item['quantity'] = float(item['quantity']) + 1
            found = True
            break
    if not found:
        cart.append({
            'product_id': product.id,
            'name': str(product),
            'barcode': product.barcode,
            'unit_price': float(product.selling_price),
            'quantity': 1,
            'purchase_cost': purchase_cost,
        })
    request.session['cart'] = cart
    return recalc_cart(request)


@login_required
@require_POST
def apply_discount(request):
    data = json.loads(request.body)
    disc_type = data.get('type', 'none')
    value = Decimal(str(data.get('value', 0)))
    request.session['discount'] = {'type': disc_type, 'value': float(value)}
    return recalc_cart(request)


@login_required
@require_POST
@transaction.atomic
def hold_order(request):
    cart = request.session.get('cart', [])
    if not cart:
        return JsonResponse({'error': 'سبد خالی است'}, status=400)
    disc = request.session.get('discount', {'type': 'none', 'value': 0})
    order = Order.objects.create(
        user=request.user,
        is_held=True,
        held_at=timezone.now(),
        discount_type=disc['type'],
        discount_value=disc['value'],
    )
    subtotal = Decimal('0')
    total_tax = Decimal('0')
    for item_data in cart:
        product = Product.objects.get(pk=item_data['product_id'])
        qty = _parse_qty(item_data['quantity'])
        price = Decimal(str(item_data['unit_price']))
        tax, _ = calculate_item_tax(price, qty, product)
        avg_cost = _get_avg_cost(product)

        OrderItem.objects.create(
            order=order, product=product, quantity=qty, unit_price=price,
            tax_amount=tax, purchase_cost=avg_cost
        )
        subtotal += price * qty
        total_tax += tax

    before = subtotal + total_tax
    disc_amount = Decimal('0')
    if disc['type'] == 'percent':
        disc_amount = (before * Decimal(str(disc['value']))) / Decimal('100')
    elif disc['type'] == 'amount':
        disc_amount = Decimal(str(disc['value']))
    order.subtotal = subtotal
    order.total_tax = total_tax
    order.discount_amount = disc_amount
    order.grand_total = before - disc_amount
    order.save()
    request.session['cart'] = []
    request.session['discount'] = {'type': 'none', 'value': 0}
    return JsonResponse({'order_id': order.id, 'message': 'سبد نگهداری شد'})


@login_required
def list_held_orders(request):
    held = Order.objects.filter(is_held=True).order_by('-held_at')[:10]
    data = []
    for o in held:
        local_time = timezone.localtime(o.held_at)
        j_date = jdatetime.datetime.fromgregorian(datetime=local_time)
        data.append({
            'id': o.id,
            'grand_total': float(o.grand_total),
            'items_count': o.items.count(),
            'held_at': j_date.strftime('%Y/%m/%d %H:%M')
        })
    return JsonResponse(data, safe=False)


@login_required
@require_POST
def resume_held_order(request):
    data = json.loads(request.body)
    order_id = data.get('order_id')
    order = get_object_or_404(Order, pk=order_id, is_held=True)
    cart = []
    for item in order.items.all():
        cart.append({
            'product_id': item.product_id,
            'name': str(item.product),
            'barcode': item.product.barcode,
            'unit_price': float(item.unit_price),
            'quantity': float(item.quantity),
            'purchase_cost': float(item.purchase_cost),
        })
    request.session['cart'] = cart
    request.session['discount'] = {
        'type': order.discount_type,
        'value': float(order.discount_value)
    }
    order.delete()
    return recalc_cart(request)


@login_required
@require_POST
def delete_held_order(request):
    data = json.loads(request.body)
    order_id = data.get('order_id')
    order = get_object_or_404(Order, pk=order_id, is_held=True)
    order.delete()
    return JsonResponse({'message': 'سبد نگهداری‌شده حذف شد.'})


@login_required
@require_POST
@transaction.atomic
def checkout(request):
    cart = request.session.get('cart', [])
    if not cart:
        return JsonResponse({'error': 'سبد خالی است'}, status=400)

    data = json.loads(request.body)
    cash_amount = Decimal(str(data.get('cash_amount', 0)))
    card_amount = Decimal(str(data.get('card_amount', 0)))
    credit_amount = Decimal(str(data.get('credit_amount', 0)))
    customer_id = data.get('customer_id')
    discount = request.session.get('discount', {'type': 'none', 'value': 0})

    product_ids = [item['product_id'] for item in cart]
    products_map = {p.pk: p for p in Product.objects.select_for_update().filter(pk__in=product_ids)}

    for item_data in cart:
        product = products_map.get(item_data['product_id'])
        if not product:
            return JsonResponse({'error': f'محصول یافت نشد'}, status=400)
        try:
            qty = _parse_qty(item_data['quantity'])
        except ValueError:
            return JsonResponse({'error': 'تعداد نامعتبر'}, status=400)
        if product.stock_quantity < qty:
            return JsonResponse(
                {'error': f'موجودی «{product}» کافی نیست. موجودی فعلی: {product.stock_quantity}'},
                status=400
            )

    subtotal = Decimal('0')
    total_tax = Decimal('0')
    for item in cart:
        product = products_map[item['product_id']]
        qty = _parse_qty(item['quantity'])
        price = Decimal(str(item['unit_price']))
        tax, _ = calculate_item_tax(price, qty, product)
        subtotal += price * qty
        total_tax += tax

    before_discount = subtotal + total_tax
    disc_amount = Decimal('0')
    if discount['type'] == 'percent':
        disc_amount = (before_discount * Decimal(str(discount['value']))) / Decimal('100')
    elif discount['type'] == 'amount':
        disc_amount = Decimal(str(discount['value']))
    grand_total = before_discount - disc_amount

    total_paid = cash_amount + card_amount + credit_amount
    if total_paid < grand_total:
        return JsonResponse({'error': 'مبلغ پرداختی کمتر از مبلغ فاکتور است.'}, status=400)

    customer = None
    if customer_id:
        try:
            customer = Customer.objects.get(pk=customer_id)
        except Customer.DoesNotExist:
            return JsonResponse({'error': 'مشتری نامعتبر'}, status=400)
        if credit_amount > 0:
            if customer.credit < credit_amount:
                return JsonResponse({'error': 'اعتبار مشتری کافی نیست'}, status=400)
            customer.credit -= credit_amount
            customer.save()

    order = Order.objects.create(
        user=request.user,
        customer=customer,
        subtotal=subtotal,
        total_tax=total_tax,
        discount_type=discount['type'],
        discount_value=discount['value'],
        discount_amount=disc_amount,
        grand_total=grand_total,
        is_paid=True
    )

    for item_data in cart:
        product = products_map[item_data['product_id']]
        qty = _parse_qty(item_data['quantity'])
        price = Decimal(str(item_data['unit_price']))
        tax, tax_lines = calculate_item_tax(price, qty, product)
        avg_cost = _get_avg_cost(product)

        order_item = OrderItem.objects.create(
            order=order,
            product=product,
            quantity=qty,
            unit_price=price,
            tax_amount=tax,
            purchase_cost=avg_cost
        )
        for line in tax_lines:
            TaxLine.objects.create(
                order_item=order_item,
                rule_name=line['rule_name'],
                amount=line['amount']
            )

        product.inventory_cost -= avg_cost * qty
        product.stock_quantity -= qty
        product.save()

        StockMovement.objects.create(
            product=product,
            quantity=qty,
            movement_type='out',
            reference=f"Order #{order.id}"
        )

    if cash_amount > 0:
        Payment.objects.create(order=order, amount=cash_amount, method='cash')
    if card_amount > 0:
        Payment.objects.create(order=order, amount=card_amount, method='card')
    if credit_amount > 0:
        Payment.objects.create(order=order, amount=credit_amount, method='credit')

    request.session['cart'] = []
    request.session['discount'] = {'type': 'none', 'value': 0}

    return JsonResponse({
        'order_id': order.id,
        'grand_total': float(grand_total),
        'message': 'فروش با موفقیت ثبت شد'
    })


@login_required
@require_POST
def update_cart_item(request):
    data = json.loads(request.body)
    index = data.get('index')
    action = data.get('action')
    cart = request.session.get('cart', [])
    if index is None or not (0 <= index < len(cart)):
        return JsonResponse({'error': 'ایندکس نامعتبر'}, status=400)
    item = cart[index]
    if action == 'increase':
        item['quantity'] = float(item['quantity']) + 1
    elif action == 'decrease':
        item['quantity'] = float(item['quantity']) - 1
        if item['quantity'] <= 0:
            cart.pop(index)
    elif action == 'set':
        try:
            new_qty = _parse_qty(data.get('quantity'))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'مقدار تعداد نامعتبر است.'}, status=400)
        item['quantity'] = float(new_qty)
    else:
        return JsonResponse({'error': 'action نامعتبر'}, status=400)
    request.session['cart'] = cart
    return recalc_cart(request)


@login_required
@require_POST
def remove_from_cart(request):
    data = json.loads(request.body)
    index = data.get('index')
    cart = request.session.get('cart', [])
    if index is not None and 0 <= index < len(cart):
        cart.pop(index)
        request.session['cart'] = cart
        return recalc_cart(request)
    return JsonResponse({'error': 'ایندکس نامعتبر'}, status=400)


# ==================== مدیریت خرید ====================

@login_required
def purchase_list(request):
    purchases = Purchase.objects.all().order_by('-created_at')
    return render(request, 'sales/purchase_list.html', {'purchases': purchases})


@login_required
@transaction.atomic
def purchase_create(request):
    if request.method == 'POST':
        supplier_id = request.POST.get('supplier')
        supplier = get_object_or_404(Supplier, pk=supplier_id)
        item_count = int(request.POST.get('item_count', 0))

        if item_count == 0:
            return JsonResponse({'status': 'error', 'error': 'حداقل یک کالا الزامی است'})

        purchase = Purchase.objects.create(supplier=supplier, total_amount=0)
        total = Decimal('0')

        for i in range(item_count):
            product_id = request.POST.get(f'product_{i}')
            qty = int(request.POST.get(f'qty_{i}', 0))
            price = Decimal(request.POST.get(f'price_{i}', 0))

            if qty <= 0 or price < 0:
                continue

            product = get_object_or_404(Product, pk=product_id)

            purchase_item = PurchaseItem.objects.create(
                purchase=purchase, product=product, quantity=qty, unit_price=price
            )
            total += price * qty

            product.inventory_cost += qty * price
            product.stock_quantity += qty
            product.purchase_price = price
            product.save()

            StockMovement.objects.create(
                product=product,
                quantity=qty,
                movement_type='in',
                reference=f"Purchase #{purchase.id} Item #{purchase_item.id}"
            )

        purchase.total_amount = total
        purchase.save()
        return JsonResponse({'status': 'ok', 'purchase_id': purchase.id})
    else:
        suppliers = Supplier.objects.all()
        return render(request, 'sales/purchase_create.html', {'suppliers': suppliers})


# ==================== گزارش‌ها و داشبورد ====================

@login_required
def dashboard(request):
    today = timezone.localdate()
    today_jalali = jdatetime.date.today().strftime('%Y/%m/%d')

    gross_sales = Order.objects.filter(
        created_at__date=today, is_paid=True, is_return=False
    ).aggregate(total=Sum(F('subtotal') - F('discount_amount')))['total'] or 0

    return_sales = Order.objects.filter(
        created_at__date=today, is_paid=True, is_return=True
    ).aggregate(total=Sum(F('subtotal') - F('discount_amount')))['total'] or 0

    today_sales = gross_sales - return_sales
    today_count = Order.objects.filter(created_at__date=today, is_paid=True, is_return=False).count()

    low_stock_products = Product.objects.filter(is_active=True, stock_quantity__lt=10).order_by('stock_quantity')[:10]
    low_stock_count = Product.objects.filter(is_active=True, stock_quantity__lt=10).count()

    today_orders = Order.objects.select_related('user')\
                     .prefetch_related('items__product')\
                     .filter(created_at__date=today, is_paid=True)\
                     .order_by('-created_at')

    cost_normal = OrderItem.objects.filter(
        order__created_at__date=today,
        order__is_paid=True,
        order__is_return=False
    ).aggregate(
        total=Sum(
            F('purchase_cost') * (F('quantity') - F('returned_quantity')),
            output_field=DecimalField(max_digits=14, decimal_places=0)
        )
    )['total'] or Decimal('0')

    cost_return = OrderItem.objects.filter(
        order__created_at__date=today,
        order__is_paid=True,
        order__is_return=True
    ).aggregate(
        total=Sum(
            F('purchase_cost') * (F('quantity') - F('returned_quantity')),
            output_field=DecimalField(max_digits=14, decimal_places=0)
        )
    )['total'] or Decimal('0')

    net_cost = cost_normal - cost_return
    profit_today = today_sales - net_cost

    total_tax = Order.objects.filter(
        created_at__date=today, is_paid=True, is_return=False
    ).aggregate(total=Sum('total_tax'))['total'] or 0

    context = {
        'today': today,
        'today_jalali': today_jalali,
        'today_sales': today_sales,
        'profit_today': profit_today,
        'today_count': today_count,
        'low_stock_products': low_stock_products,
        'low_stock_count': low_stock_count,
        'today_orders': today_orders,
        'total_tax': total_tax,
    }
    return render(request, 'sales/dashboard.html', context)


@login_required
def daily_report(request):
    start_str = request.GET.get('start_date')
    end_str = request.GET.get('end_date')

    today = timezone.localdate()

    try:
        if start_str:
            start_date = date.fromisoformat(start_str)
        else:
            start_date = today

        if end_str:
            end_date = date.fromisoformat(end_str)
        else:
            end_date = today
    except (ValueError, TypeError):
        start_date = today
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    display_start = jdatetime.date.fromgregorian(date=start_date).strftime('%Y/%m/%d')
    display_end = jdatetime.date.fromgregorian(date=end_date).strftime('%Y/%m/%d')
    input_start = start_date.isoformat()
    input_end = end_date.isoformat()

    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        is_paid=True,
        is_return=False
    ).select_related('user').prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related('product'))
    ).order_by('-created_at')

    gross_sales = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        is_paid=True,
        is_return=False
    ).aggregate(total=Sum(F('subtotal') - F('discount_amount')))['total'] or 0

    return_sales = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        is_paid=True,
        is_return=True
    ).aggregate(total=Sum(F('subtotal') - F('discount_amount')))['total'] or 0

    total_sales = gross_sales - return_sales
    total_tax = orders.aggregate(total=Sum('total_tax'))['total'] or 0
    order_count = orders.count()

    cost_normal = OrderItem.objects.filter(
        order__created_at__date__gte=start_date,
        order__created_at__date__lte=end_date,
        order__is_paid=True,
        order__is_return=False
    ).aggregate(
        total=Sum(
            F('purchase_cost') * (F('quantity') - F('returned_quantity')),
            output_field=DecimalField(max_digits=14, decimal_places=0)
        )
    )['total'] or Decimal('0')

    cost_return = OrderItem.objects.filter(
        order__created_at__date__gte=start_date,
        order__created_at__date__lte=end_date,
        order__is_paid=True,
        order__is_return=True
    ).aggregate(
        total=Sum(
            F('purchase_cost') * (F('quantity') - F('returned_quantity')),
            output_field=DecimalField(max_digits=14, decimal_places=0)
        )
    )['total'] or Decimal('0')

    total_cost = cost_normal - cost_return
    gross_profit = total_sales - total_cost

    # ── داده‌های نمودار: روزانه یا ساعتی ──
    chart_labels = []
    chart_values = []

    if start_date == end_date:
        # ── حالت تک‌روزه: فروش ساعتی ──
        from collections import OrderedDict
        hourly_qs = Order.objects.filter(
            created_at__date=start_date,
            is_paid=True
        ).annotate(
            hour=TruncHour('created_at')
        ).values('hour').annotate(
            gross=Sum(F('subtotal') - F('discount_amount'), filter=Q(is_return=False)),
            returns=Sum(F('subtotal') - F('discount_amount'), filter=Q(is_return=True))
        ).order_by('hour')

        # ساخت یک دیکشنری از ساعت (۰-۲۳) به فروش خالص
        hour_map = {h: Decimal('0') for h in range(24)}
        for entry in hourly_qs:
            h = entry['hour'].hour if entry['hour'] else 0
            net = (entry['gross'] or 0) - (entry['returns'] or 0)
            hour_map[h] += net

        # تولید برچسب فارسی برای هر ساعت (فقط ساعاتی که فروش دارند یا همیشه ۰ تا ۲۳)
        # اما بهتر است همه ۲۴ ساعت را نشان دهیم تا روند صفرها هم مشخص باشد
        for h in range(24):
            chart_labels.append(f"{h:02d}:00")   # می‌توان با jdatetime شمسی کرد ولی همین کافی است
            chart_values.append(float(hour_map[h]))
    else:
        # ── حالت بازه‌ای: فروش روزانه ──
        daily_sales_qs = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_paid=True
        ).annotate(
            day=TruncDate('created_at')
        ).values('day').annotate(
            gross=Sum(F('subtotal') - F('discount_amount'), filter=Q(is_return=False)),
            returns=Sum(F('subtotal') - F('discount_amount'), filter=Q(is_return=True))
        ).order_by('day')

        for entry in daily_sales_qs:
            gregorian_date = entry['day']
            if gregorian_date:
                net = (entry['gross'] or 0) - (entry['returns'] or 0)
                j_date = jdatetime.date.fromgregorian(date=gregorian_date)
                chart_labels.append(j_date.strftime('%Y/%m/%d'))
                chart_values.append(float(net))

    # ── پردازش آیتم‌های فروش به تفکیک کالا ──
    normal_items = OrderItem.objects.filter(
        order__created_at__date__gte=start_date,
        order__created_at__date__lte=end_date,
        order__is_paid=True,
        order__is_return=False
    ).select_related('product')

    return_items = OrderItem.objects.filter(
        order__created_at__date__gte=start_date,
        order__created_at__date__lte=end_date,
        order__is_paid=True,
        order__is_return=True
    ).select_related('product')

    product_sales = {}
    for item in normal_items:
        net_qty = item.quantity - item.returned_quantity
        if net_qty <= 0:
            continue
        prod_id = item.product_id
        if prod_id not in product_sales:
            product_sales[prod_id] = {
                'product': item.product,
                'quantity': Decimal('0'),
                'total_sales': Decimal('0'),
                'total_cost': Decimal('0'),
            }
        product_sales[prod_id]['quantity'] += net_qty
        tax_share = item.tax_amount * (net_qty / item.quantity) if item.quantity else Decimal('0')
        product_sales[prod_id]['total_sales'] += item.unit_price * net_qty + tax_share
        product_sales[prod_id]['total_cost'] += item.purchase_cost * net_qty

    for item in return_items:
        net_qty = item.quantity - item.returned_quantity
        if net_qty <= 0:
            continue
        prod_id = item.product_id
        if prod_id not in product_sales:
            product_sales[prod_id] = {
                'product': item.product,
                'quantity': Decimal('0'),
                'total_sales': Decimal('0'),
                'total_cost': Decimal('0'),
            }
        product_sales[prod_id]['quantity'] -= net_qty
        tax_share = item.tax_amount * (net_qty / item.quantity) if item.quantity else Decimal('0')
        product_sales[prod_id]['total_sales'] -= item.unit_price * net_qty + tax_share
        product_sales[prod_id]['total_cost'] -= item.purchase_cost * net_qty

    for data in product_sales.values():
        data['profit'] = data['total_sales'] - data['total_cost']

    product_sales_list = sorted(product_sales.values(), key=lambda x: x['quantity'], reverse=True)

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'display_start': display_start,
        'display_end': display_end,
        'input_start': input_start,
        'input_end': input_end,
        'orders': orders,
        'total_sales': total_sales,
        'total_tax': total_tax,
        'order_count': order_count,
        'gross_profit': gross_profit,
        'product_sales': product_sales_list,
        'chart_labels_json': json.dumps(chart_labels, ensure_ascii=False),
        'chart_values_json': json.dumps(chart_values),
    }
    return render(request, 'sales/daily_report.html', context)



@login_required
def detailed_tax_report(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    if start_date and end_date:
        tax_data = TaxLine.objects.filter(
            order_item__order__created_at__date__range=[start_date, end_date]
        ).values('rule_name').annotate(total=Sum('amount')).order_by('rule_name')
    else:
        tax_data = TaxLine.objects.values('rule_name').annotate(total=Sum('amount')).order_by('rule_name')
    context = {'tax_data': tax_data, 'start': start_date, 'end': end_date}
    return render(request, 'sales/tax_report.html', context)


@login_required
def receipt_view(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    return render(request, 'sales/receipt.html', {'order': order})


@login_required
def invoice_xml(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<Invoice>
    <Id>{order.id}</Id>
    <Date>{order.created_at.strftime('%Y-%m-%d')}</Date>
    <Total>{order.grand_total}</Total>
    <Tax>{order.total_tax}</Tax>
</Invoice>"""
    response = HttpResponse(xml_content, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="invoice_{order.id}.xml"'
    return response


# ==================== ورودی/خروجی اکسل ====================

MODEL_ORDER = [
    'inventory.Category',
    'inventory.Supplier',
    'inventory.Product',
    'inventory.StockMovement',
    'inventory.Purchase',
    'inventory.PurchaseItem',
    'taxation.TaxRule',
    'taxation.ProductTaxProfile',
    'taxation.TaxLine',
    'sales.Order',
    'sales.OrderItem',
    'sales.Payment',
]

def get_ordered_models():
    result = []
    for label in MODEL_ORDER:
        app_label, model_name = label.split('.')
        result.append(apps.get_model(app_label, model_name))
    return result

@staff_member_required
def export_data(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for model in get_ordered_models():
        sheet_name = f"{model._meta.app_label}.{model._meta.model_name}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        fields = [f for f in model._meta.get_fields() if f.concrete and not f.many_to_many]
        headers = [f.verbose_name if f.verbose_name else f.name for f in fields]
        ws.append(headers)
        for obj in model.objects.all():
            row = []
            for field in fields:
                value = getattr(obj, field.attname if hasattr(field, 'attname') else field.name)
                if isinstance(value, Decimal):
                    value = float(value)
                elif isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, models.Model):
                    value = value.pk
                row.append(value)
            ws.append(row)
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(header)) + 2, 12)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="database_backup.xlsx"'
    return response

@staff_member_required
@transaction.atomic
def import_data(request):
    if request.method == 'POST' and request.FILES.get('data_file'):
        excel_file = request.FILES['data_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "فقط فایل‌های xlsx مجاز هستند.")
            return redirect('import_export')
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            messages.error(request, f"خطا در باز کردن فایل: {e}")
            return redirect('import_export')
        errors = []
        success_count = 0
        for model in get_ordered_models():
            sheet_name = f"{model._meta.app_label}.{model._meta.model_name}"[:31]
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            field_map = {}
            all_fields = {f.name: f for f in model._meta.get_fields() if f.concrete and not f.many_to_many}
            verbose_to_name = {}
            for f in all_fields.values():
                vname = f.verbose_name if f.verbose_name else f.name
                verbose_to_name[vname] = f.name
            for col_idx, header in enumerate(headers):
                if header in verbose_to_name:
                    field_map[col_idx] = verbose_to_name[header]
                elif header in all_fields:
                    field_map[col_idx] = header
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                data = {}
                pk = None
                for col_idx, field_name in field_map.items():
                    value = row[col_idx]
                    if field_name == 'id':
                        pk = value
                        continue
                    field = all_fields.get(field_name)
                    if not field:
                        continue
                    try:
                        if isinstance(field, models.ForeignKey):
                            if value is not None:
                                related_model = field.related_model
                                if not related_model.objects.filter(pk=value).exists():
                                    raise ValueError(f"شناسه {value} برای {related_model.__name__} یافت نشد.")
                                value = related_model.objects.get(pk=value)
                            else:
                                value = None
                        elif isinstance(field, models.DecimalField):
                            if value is not None:
                                value = Decimal(str(value))
                        elif isinstance(field, models.DateTimeField):
                            if value is not None:
                                value = datetime.fromisoformat(str(value))
                        elif isinstance(field, models.BooleanField):
                            if value is not None:
                                value = bool(value)
                    except Exception as e:
                        errors.append(f"خطا در {model.__name__}، فیلد {field_name}: {e}")
                        continue
                    data[field_name] = value
                if errors:
                    continue
                try:
                    if pk is not None:
                        model.objects.update_or_create(id=pk, defaults=data)
                    else:
                        model.objects.create(**data)
                    success_count += 1
                except Exception as e:
                    errors.append(f"خطا در ذخیره {model.__name__}: {e}")
        if errors:
            messages.warning(request, f"{len(errors)} خطا رخ داد. اولین خطا: {errors[0]}")
        else:
            messages.success(request, f"{success_count} رکورد با موفقیت وارد شد.")
        return redirect('import_export')
    else:
        return render(request, 'sales/import_export.html')


# ==================== برگشت از فروش ====================

@login_required
@require_POST
@transaction.atomic
def return_order(request, order_id):
    original_order = get_object_or_404(Order, pk=order_id, is_paid=True, is_return=False)

    if Order.objects.filter(original_order=original_order, is_return=True).exists():
        messages.error(request, 'این فاکتور قبلاً برگشت خورده است.')
        return redirect('management:order_list')

    return_order_obj = Order.objects.create(
        user=request.user,
        customer=original_order.customer,
        subtotal=original_order.subtotal,
        total_tax=original_order.total_tax,
        discount_type=original_order.discount_type,
        discount_value=original_order.discount_value,
        discount_amount=original_order.discount_amount,
        grand_total=original_order.grand_total,
        is_paid=True,
        is_return=True,
        original_order=original_order
    )

    for item in original_order.items.all():
        product = item.product
        qty = item.quantity
        purchase_cost = item.purchase_cost

        product.inventory_cost += purchase_cost * qty
        product.stock_quantity += qty
        product.save()

        StockMovement.objects.create(
            product=product,
            quantity=qty,
            movement_type='in',
            reference=f"Return from Order #{original_order.id}"
        )

        new_item = OrderItem.objects.create(
            order=return_order_obj,
            product=product,
            quantity=qty,
            unit_price=item.unit_price,
            tax_amount=item.tax_amount,
            purchase_cost=purchase_cost
        )
        for tax_line in item.tax_lines.all():
            TaxLine.objects.create(
                order_item=new_item,
                rule_name=tax_line.rule_name,
                amount=tax_line.amount
            )

    for payment in original_order.payment_set.all():
        Payment.objects.create(
            order=return_order_obj,
            amount=-payment.amount,
            method=payment.method,
            paid_at=timezone.now()
        )

    messages.success(request, 'فاکتور با موفقیت برگشت خورد.')
    return redirect('management:order_list')


# ==================== اصلاحیه سفارش (با تخفیف نسبی) ====================

@login_required
@transaction.atomic
def edit_order(request, order_id):
    original_order = get_object_or_404(Order, pk=order_id, is_paid=True, is_return=False)

    if request.method == 'POST':
        return_items = request.POST.getlist('return_item_ids[]')
        return_quantities = request.POST.getlist('return_quantities[]')
        payment_method = request.POST.get('payment_method', 'cash')

        total_return_value = Decimal('0')       # مبلغ کامل برگشتی‌ها (قبل از تخفیف)
        total_return_tax = Decimal('0')
        total_return_discount = Decimal('0')    # جمع تخفیف برگشتی

        # ۱. محاسبه نسبت تخفیف
        before_discount = original_order.subtotal + original_order.total_tax
        discount_ratio = Decimal('0')
        if before_discount > 0 and original_order.discount_amount > 0:
            discount_ratio = original_order.discount_amount / before_discount

        # ۲. پردازش اقلام برگشتی
        for item_id_str, qty_str in zip(return_items, return_quantities):
            try:
                qty_to_return = Decimal(str(qty_str))
            except (ValueError, TypeError):
                continue
            if qty_to_return <= 0:
                continue

            original_item = get_object_or_404(OrderItem, pk=int(item_id_str), order=original_order)
            available = original_item.quantity - original_item.returned_quantity
            if qty_to_return > available:
                messages.error(request, f'مقدار برگشتی برای {original_item.product} نمی‌تواند بیشتر از {available} باشد.')
                return redirect('edit_order', order_id=order_id)

            product = original_item.product

            # موجودی و ارزش انبار
            product.inventory_cost += original_item.purchase_cost * qty_to_return
            product.stock_quantity += qty_to_return
            product.save()

            StockMovement.objects.create(
                product=product,
                quantity=qty_to_return,
                movement_type='in',
                reference=f"Edit Order #{original_order.id}"
            )

            # به‌روزرسانی quantities در OrderItem
            original_item.returned_quantity += qty_to_return
            if original_item.returned_quantity >= original_item.quantity:
                original_item.is_returned = True
            original_item.save()

            # محاسبه مبلغ کامل این برگشتی
            tax_share = original_item.tax_amount * (qty_to_return / original_item.quantity) if original_item.quantity else Decimal('0')
            item_full_value = original_item.unit_price * qty_to_return + tax_share

            # سهم تخفیف این برگشتی
            item_discount = item_full_value * discount_ratio if discount_ratio else Decimal('0')

            total_return_value += item_full_value
            total_return_tax += tax_share
            total_return_discount += item_discount

        # ۳. پردازش اقلام جدید (بدون تخفیف)
        new_product_ids = request.POST.getlist('new_product_ids[]')
        new_quantities = request.POST.getlist('new_quantities[]')

        total_new_value = Decimal('0')
        total_new_tax = Decimal('0')

        for prod_id_str, qty_str in zip(new_product_ids, new_quantities):
            if not prod_id_str:
                continue
            try:
                qty = _parse_qty(qty_str)
            except (ValueError, TypeError):
                continue

            product = get_object_or_404(Product, pk=int(prod_id_str), is_active=True)

            if product.stock_quantity < qty:
                messages.error(request, f'موجودی «{product}» کافی نیست. موجودی فعلی: {product.stock_quantity}')
                return redirect('edit_order', order_id=order_id)

            price = product.selling_price
            tax, tax_lines = calculate_item_tax(price, qty, product)
            avg_cost = _get_avg_cost(product)

            # کاهش موجودی
            product.inventory_cost -= avg_cost * qty
            product.stock_quantity -= qty
            product.save()

            StockMovement.objects.create(
                product=product,
                quantity=qty,
                movement_type='out',
                reference=f"Edit Order #{original_order.id}"
            )

            new_item = OrderItem.objects.create(
                order=original_order,
                product=product,
                quantity=qty,
                unit_price=price,
                tax_amount=tax,
                purchase_cost=avg_cost,
            )
            for line in tax_lines:
                TaxLine.objects.create(
                    order_item=new_item,
                    rule_name=line['rule_name'],
                    amount=line['amount'],
                )

            total_new_value += price * qty + tax
            total_new_tax += tax

        # ۴. محاسبه مابه‌التفاوت واقعی (با احتساب تخفیف برگشتی)
        net_return_value = total_return_value - total_return_discount
        value_diff = total_new_value - net_return_value

        if value_diff > 0:
            Payment.objects.create(
                order=original_order,
                amount=value_diff,
                method=payment_method,
                paid_at=timezone.now(),
            )
            messages.info(request, f'مشتری باید {value_diff:,.0f} تومان پرداخت کند.')
        elif value_diff < 0:
            Payment.objects.create(
                order=original_order,
                amount=value_diff,
                method=payment_method,
                paid_at=timezone.now(),
            )
            messages.info(request, f'باید {-value_diff:,.0f} تومان به مشتری برگشت داده شود.')
        else:
            messages.success(request, 'اصلاحیه بدون تغییر مبلغ ثبت شد.')

        # ۵. به‌روزرسانی فاکتور اصلی
        original_order.subtotal = original_order.subtotal - (total_return_value - total_return_tax) + (total_new_value - total_new_tax)
        original_order.total_tax = original_order.total_tax - total_return_tax + total_new_tax
        original_order.discount_amount = original_order.discount_amount - total_return_discount
        original_order.grand_total = original_order.subtotal + original_order.total_tax - original_order.discount_amount
        original_order.save()

        return redirect('management:order_detail', pk=original_order.id)

    # GET: نمایش فرم
    products = Product.objects.filter(is_active=True)
    order_items = []
    for item in original_order.items.all():
        item.available_qty = item.quantity - item.returned_quantity
        order_items.append(item)

    return render(request, 'sales/edit_order.html', {
        'order': original_order,
        'order_items': order_items,
        'products': products,
    })


@login_required
def search_products_for_purchase(request):
    query = request.GET.get('q', '')
    products = Product.objects.filter(is_active=True).filter(
        Q(name__icontains=query) | Q(barcode__icontains=query)
    )[:20]
    results = [{
        'id': p.id,
        'name': str(p),
        'barcode': p.barcode,
        'stock': float(p.stock_quantity),
    } for p in products]
    return JsonResponse(results, safe=False)


@login_required
def low_stock_report(request):
    category_id = request.GET.get('category')
    try:
        min_qty = Decimal(request.GET.get('min_qty', '5'))
    except Exception:
        min_qty = Decimal('5')

    products = Product.objects.filter(is_active=True, stock_quantity__lt=min_qty)

    selected_category = None
    if category_id:
        try:
            selected_category = Category.objects.get(pk=category_id)
            products = products.filter(
                Q(category=selected_category) | Q(category__parent=selected_category)
            )
        except Category.DoesNotExist:
            pass

    products = products.select_related('category').order_by('stock_quantity')

    for p in products:
        if p.stock_quantity > 0:
            p.avg_cost = p.inventory_cost / p.stock_quantity
            p.total_value = p.inventory_cost
        else:
            p.avg_cost = Decimal('0')
            p.total_value = Decimal('0')

    context = {
        'products': products,
        'categories': Category.objects.all(),
        'selected_category': selected_category,
        'min_qty': min_qty,
    }
    return render(request, 'sales/low_stock_report.html', context)